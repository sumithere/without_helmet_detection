from ultralytics import YOLO
import cv2
import pytesseract
import os
import sqlite3
import openpyxl
import random
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import time
import math


# Function to create the table if it doesn't exist
def create_table():
    # Connect to the SQLite database
    conn = sqlite3.connect("vehicle_data.db")
    # Create a cursor object to execute SQL statements
    c = conn.cursor()
    # Create the "vehicles" table with two columns: "vehicle_number" and "bike_image_path"
    # The "IF NOT EXISTS" clause ensures that the table is only created if it doesn't already exist
    c.execute("CREATE TABLE IF NOT EXISTS vehicles (vehicle_number TEXT, bike_image_path TEXT)")
    # Commit the changes to the database
    conn.commit()
    # Close the database connection
    conn.close()

# Function to insert a record into the "vehicles" table
def insert_record(vehicle_number, bike_image_path):
    # Connect to the SQLite database
    conn = sqlite3.connect("vehicle_data.db")
    # Create a cursor object to execute SQL statements
    c = conn.cursor()
    # Insert the record into the "vehicles" table using parameterized SQL statement
    c.execute("INSERT INTO vehicles VALUES (?, ?)", (vehicle_number, bike_image_path))
    # Commit the changes to the database
    conn.commit()
    # Close the database connection
    conn.close()

curr_row=2


def create_excel_sheet(filename, sheet_name , data):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    
    # Create a new sheet
    sheet = workbook.active
    sheet.title = sheet_name
    
    for col_idx, col_value in enumerate(data, start=1):
        sheet.cell(row=1, column=col_idx, value=col_value)
   
    # Save the workbook
    workbook.save(filename)
    print(f"Excel sheet '{sheet_name}' created successfully with data.")
    

def insert_data_into_excel(filename, sheet_name, data, start_row=1, start_col=1):
    try:
        global curr_row
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)
        
        # Select the specified sheet
        sheet = workbook[sheet_name]
        
        curr_row+=1

        # Insert data into the sheet starting from the specified row and column
        for col_idx, col_value in enumerate(data, start=start_col):
            sheet.cell(row=start_row, column=col_idx, value=col_value)
        print(curr_row)
        # Save the changes
        workbook.save(filename)
        print(f"Data inserted successfully into '{sheet_name}' in {filename}.")
    except FileNotFoundError:
        print("File not found. Please provide a valid filename.")
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in {filename}.")



# Initialize YOLO models
person_bike_model = YOLO("C:\\Users\\sumit\\iotProject\\Real-Time-Helmet-Detection-and-License-Number-Extraction-for-Traffic-Rule-Enforcement\\runsOfAllThreeModels\\bike\\runs\\detect\\train2\\weights\\best.pt")
helmet_model = YOLO("C:\\Users\\sumit\\iotProject\\Real-Time-Helmet-Detection-and-License-Number-Extraction-for-Traffic-Rule-Enforcement\\runsOfAllThreeModels\\helmet\\runs\\detect\\train10\\weights\\best.pt")
number_plate_model = YOLO("C:\\Users\\sumit\\iotProject\\Real-Time-Helmet-Detection-and-License-Number-Extraction-for-Traffic-Rule-Enforcement\\runsOfAllThreeModels\\number_plate\\runs\\detect\\train4\\weights\\best.pt")

# Set up Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"  # Update with the path to your Tesseract OCR executable

output_dir = "C:\\Users\\sumit\\iotProject\\Real-Time-Helmet-Detection-and-License-Number-Extraction-for-Traffic-Rule-Enforcement\\test_output_iotProject"  # Directory to save the output images
# Set up video capture

video_path = "C:\\Users\\sumit\\Downloads\\without_helmet_video2.mp4"
# Set up video capture for a file or camera
use_camera = False  # Set this to True if using a live camera feed

# Capture from file or camera based on flag
if use_camera:
    video_capture = cv2.VideoCapture(0)  # Use 0 for default camera
else:
    video_capture = cv2.VideoCapture(video_path)  # Use path for video file


if not video_capture.isOpened():
    print("Error: Could not open the video file.")
else:
    while video_capture.isOpened():
        # Capture frame-by-frame
        ret, frame = video_capture.read()
        # Process frame
        print(f"ret: {ret}, frame type: {type(frame)}")
        if not ret:
            if use_camera:
                print("Error reading camera feed.")
            else:
                print("End of video or error reading frame.")
            break
        
        
        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        
        
        # img0 = 'C:\\Users\\sumit\\Downloads\\without_helmet_images\\432.jpg'
        # image = cv2.imread(img0)
        
        
        
        
        
        
        
        
        
        
        # img0 = 'C:\\Users\\sumit\\Downloads\\without_helmet1.jpg'

        # Run YOLOv8 inference on the frame
        # results = person_bike_model(frame)

        # # Visualize the results on the frame
        # annotated_frame = results[0].plot()

        # cv2.imshow('Video Frame', annotated_frame)
        
        # if cv2.waitKey(1) & 0xFF == ord('q'):
        #     break
        # Detect person on a bike
        person_bike_results = person_bike_model.predict(img)

        # Process each detection result
        for r in person_bike_results:
            boxes = r.boxes
            # Filter detections for person on a bike
            for box in boxes:
                cls = box.cls
                # print(person_bike_model.names[int(cls)], person_bike_model.names[int(cls)] == "Person_Bikes")
                if person_bike_model.names[int(cls)] == "Person_Bike":
                    # Crop person on a bike image
                    x1, y1, x2, y2 = box.xyxy[0]
                    person_bike_image = frame[int(y1):int(y2), int(x1):int(x2)]
                    
                    
                    image_file=random.random()*100
                    output_file = f"person_violation_{image_file}"
                    output_path = os.path.join(output_dir, output_file) 
                    cv2.imwrite(output_path, person_bike_image)
                    
                    
                    # Perform OCR on the number plate image
                    # print("line128==========================================================")
                    # Detect helmet on the person
                    helmet_results = helmet_model.predict(person_bike_image)

                    # Process each helmet detection result
                    for hr in helmet_results:
                        h_boxes = hr.boxes
                        # Filter detections for no helmet
                        for h_bo in h_boxes:
                            h_cls = h_bo.cls
                            if not helmet_model.names[int(h_cls)] == "With Helmet" :
                                # Extract number plate from the person bike image
                                number_plate_results = number_plate_model.predict(person_bike_image)
                                # print("line141==============================================")
                                
                                # Process each number plate detection result
                                for npr in number_plate_results:
                                    np_boxes = npr.boxes
                                    # Filter detections for number plate
                                    for np_box in np_boxes:
                                        np_cls = np_box.cls
                                        print(number_plate_model.names[int(np_cls)])
                                        # print("line149=====================================")
                                        if number_plate_model.names[int(np_cls)] == "License_Plate":
                                            # Crop number plate image
                                            np_x1, np_y1, np_x2, np_y2 = np_box.xyxy[0]
                                            number_plate_image = person_bike_image[int(np_y1):int(np_y2),
                                                                int(np_x1):int(np_x2)]
                                            
                                            # print("line156=====================================")
                                            # Save the cropped number plate image
                                            image_file=math.trunc(random.random()*1000)
                                            output_file = f"person_violation_{image_file}"
                                            output_path = os.path.join(output_dir, output_file)
                                            cv2.imwrite(output_path, person_bike_image)
                                            # cv2.imshow('Image', person_bike_image)
                                    
                                #creating plot for person on bike        
                                            # plt.figure("Person on Bike Without Helmet")
                                            # plt.imshow(person_bike_image)
                                            # plt.axis('off')  # Turn off axis numbers and ticks
                                            # plt.show()
                                            
                                            
                                            # plt.pause(1)
                                            # plt.close()
                                        
                                # creating plot of number plate
                                            # plt.figure("Number Plate")
                                            # plt.imshow(number_plate_image)
                                            # plt.axis('off')  # Turn off axis numbers and ticks
                                            # plt.show()
                                            
                                            
                                            # plt.pause(1)
                                            # plt.close()
                                            
                                            
                                            # Perform OCR on the number plate image
                                            gray = cv2.cvtColor(number_plate_image, cv2.COLOR_BGR2GRAY)
                                            blur = cv2.GaussianBlur(gray, (3,3), 0)
                                            thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

                                            # Morph open to remove noise and invert image
                                            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
                                            opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
                                            invert = 255 - opening
                                                # Use Tesseract OCR to extract text with specified language
                                            text  = pytesseract.image_to_string(invert, lang='eng', config='--psm 6')
                                            text=text.strip()
                                            # text = pytesseract.image_to_string(gray)
                                            # Example usage
                                            # Create the "vehicles" table if it doesn't exist
                                            # create_table()
                                            # Calling the function to create an Excel sheet with the provided data
                                            
                                #create excel sheet
                                            # excel_headings =  ["Vehicle Number", "Bike Image Path"]
                                            # create_excel_sheet("without_helmet.xlsx", "Sheet1" , excel_headings)
                                            # print("excel created")
                                            
                                            
                                            # Insert two records into the "vehicles" table
                                            # insert_record(text, output_path)
                                            # insert data into excel
                                            new_data = [text, output_path]
                                            insert_data_into_excel("without_helmet.xlsx", "Sheet1", new_data, start_row=curr_row, start_col=1)
                                            # Print the extracted text
                                            print("Number Plate Text:", text)

video_capture.release()
cv2.destroyAllWindows()
